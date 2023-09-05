#This code will copy all the data of one excel sheet in another sheet

import openpyxl
#import pandas as pd

def copy_data():
    path1="C:\\Users\\158410\\OneDrive - Arrow Electronics, Inc\\Documents\\Sample1.xlsx"
    workbook= openpyxl.load_workbook(path1)
    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Sheet2"]
    maxr = sheet1.max_row
    maxc = sheet1.max_column
    for i in range(1, maxr + 1):
        for j in range(1, maxc + 1):
            print(sheet1.cell(row=i, column=j).value)
            sheet2.cell(row=i, column=j).value = sheet1.cell(row=i, column=j).value
    workbook.save("C:\\Users\\158410\\OneDrive - Arrow Electronics, Inc\\Documents\\Sample1.xlsx")

copy_data()

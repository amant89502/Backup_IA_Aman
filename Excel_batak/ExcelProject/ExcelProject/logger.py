# import inspect
# import logging
#
# class BaseClass:
#     def loggingDemo(self):
#         loggerName = inspect.stack()[1][3]
#         logger = logging.getLogger(loggerName)
#         fileHandler = logging.FileHandler('logfile.log')
#         formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
#         fileHandler.setFormatter(formatter)
#         logger.addHandler(fileHandler)  # filehandler object
#         logger.setLevel(logging.DEBUG)
#         return logger


import openpyxl

# Load the source Excel file
source_file_path = "C:/Users/158173/OneDrive - Arrow Electronics, Inc/Documents/aman_excel1.xlsx"
source_workbook = openpyxl.load_workbook(source_file_path)
source_sheet = source_workbook.active

# Create a new Excel workbook for the destination
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active

# Copy data from source sheet to destination sheet
for row in source_sheet.iter_rows(values_only=True):
    destination_sheet.append(row)

# Save the destination workbook to a new Excel file
destination_file_path = "C:/Users/158173/OneDrive - Arrow Electronics, Inc/Documents/amanexcel2.xlsx"
destination_workbook.save(destination_file_path)

# Close both workbooks
source_workbook.close()
destination_workbook.close()

print("Successfull copied !")



